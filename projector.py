"""
magiclib / projector

------------------------------------------------------------------------------------------------------------------------
magiclib / projector is a visualization and analysis module designed to handle diverse experimental and
characterization data in cultural heritage and materials science. It provides standardized functions—such as read,
plot, magic_plot, and save_json—across multiple techniques including tensile, compression, torsion, DIL, XRD, Raman,
XPS, Mapping, and XRF. In addition to general plotting, it offers specialized capabilities like PDF matching in XRD,
curve fitting in Raman and XPS, derivative plotting in DIL, and elemental analysis in Mapping and XRF. By unifying
heterogeneous workflows into a consistent structure, projector makes it easier for researchers to visualize, interpret,
and preserve data in a reproducible way.
------------------------------------------------------------------------------------------------------------------------

Xinyuan Su
"""


from . import general

import os
import io
import re
import copy
import math
import time
import json
import shutil
import inspect
import chardet
import pytesseract
import pandas as pd
import numpy as np
import seaborn as sns
from pathlib import Path
from collections import Counter
from docx import Document
from PIL import Image, ImageEnhance
from openpyxl.styles import Alignment
from tqdm import tqdm
from tabulate import tabulate
from scipy.signal import find_peaks
from prettytable import PrettyTable
from typing import Union, Tuple, Optional, List, Dict
from matplotlib import pyplot as plt
from pandas import DataFrame
from abc import ABC, abstractmethod
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from matplotlib.gridspec import GridSpec


""" 标准定义 """
class Keyword(ABC):

    # 使用全局变量作为类属性的默认值
    Magic_Database = general.Magic_Database
    Standard_Database = general.Standard_Database
    interval_time = general.interval_time

    # 读取原生文件
    @abstractmethod
    def read(self) -> Dict[str, DataFrame]:
        pass

    # 绘图
    @abstractmethod
    def plot(self, save_path: Union[bool, str] = True, **kwargs) -> None:
        pass

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    @abstractmethod
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    @abstractmethod
    def save_json(self):
        pass


""" Tensile """
class Tensile(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='tensile', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取拉力测试的源生 EXCEL 文件，如果文件的扩展名为 'xls'，则先改为 'xlsx'
        Read the raw EXCEL file of the tensile test; if the file extension is 'xls', change it to 'xlsx' first.

        针对仪器： 美国 INSTRON 生产的万能材料试验机

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        # 更改 Excel 的扩展名，若无需要更改的也不影响，只有为路径为目录的情况下才会更改
        if os.path.isdir(self.read_path):
            self.data.rename_extension(path=self.read_path, old_extension='xls', new_extension='xlsx', show=False)

        self.data.read_excel(excel_path=self.read_path)  # 读取 Excel 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, **kwargs) -> None:
        """
        绘制拉力测试的图像
        Plot the tensile test image.

        :param save_path: (str) 文件的保存路径
        :param kwargs: 可加的关键词参数

        :return: None
        """

        self.data.plot_line(save_path=save_path, **kwargs)  # 进行图形的绘制

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass


""" Compression """
class Compression(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='compression', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取压力测试的源生 EXCEL 文件，如果文件的扩展名为 'xls'，则先改为 'xlsx'
        Read the raw EXCEL file of the compression test; if the file extension is 'xls', change it to 'xlsx' first.

        针对仪器： 美国 INSTRON 生产的万能材料试验机

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        # 更改 Excel 的扩展名，若无需要更改的也不影响，只有为路径为目录的情况下才会更改
        if os.path.isdir(self.read_path):
            self.data.rename_extension(path=self.read_path, old_extension='xls', new_extension='xlsx', show=False)

        self.data.read_excel(excel_path=self.read_path)  # 读取 Excel 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, **kwargs) -> None:
        """
        绘制压力测试的图像
        Plot the compression test image.

        :param save_path: (str) 文件的保存路径
        :param kwargs: 可加的关键词参数

        :return: None
        """

        self.data.plot_line(save_path=save_path, **kwargs)  # 进行图形的绘制

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass


""" Torsion """
class Torsion(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='torsion', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取扭转测试的源生 EXCEL 文件，如果文件的扩展名为 'xls'，则先改为 'xlsx'
        Read the raw EXCEL file of the torsion test; if the file extension is 'xls', change it to 'xlsx' first.

        针对仪器： 美国 INSTRON 生产的万能材料试验机

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        # 更改 Excel 的扩展名，若无需要更改的也不影响，只有为路径为目录的情况下才会更改
        if os.path.isdir(self.read_path):
            self.data.rename_extension(path=self.read_path, old_extension='xls', new_extension='xlsx', show=False)

        self.data.read_excel(excel_path=self.read_path)  # 读取 Excel 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, show_peak: bool = True, **kwargs) -> None:
        """
        绘制扭转测试的图像
        Plot the tensile test image.

        :param save_path: (str) 文件的保存路径
        :param show_peak: (bool) 打印最大值坐标，默认为 True
        :param kwargs: 可加的关键词参数

        :return: None
        """

        self.data.plot_line(save_path=save_path, **kwargs)  # 进行图形的绘制

        if show_peak:
            self.print_peaks()

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 打印峰值
    def print_peaks(self, show_details: bool = True) -> str:
        """
        打印并返回峰值信息
        Print and return peak information

        :param show_details: (bool) 是否打印，默认为 True

        :return peak_details: (str) 返回的峰值信息
        """

        peak_details = ''
        for title, data_df in self.data.data_dic.items():
            # 找到最大值的索引
            max_index = data_df.iloc[:, 1].idxmax()

            # 获取最大峰值对应的横坐标的值
            peak_x = data_df.iloc[max_index, 0]  # 第一列的值在 max_index 的位置

            # 获取最大峰值对应的纵坐标的值
            peak_y = data_df.iloc[max_index, 1]  # 第二列的值在 max_index 的位置

            details = (f"In \033[92m{title:<20}\033[0m, "
                       f"The peak occurs at \033[94m{peak_x:4.2f}mm\033[0m "
                       f"with a value of \033[95m{peak_y:6.2f}N\033[0m")

            if show_details:
                print(details)

            peak_details += details + '\n'

        return peak_details


""" XRD """
class XRD(Keyword):
    """
    默认参数：

    测试区间为 5° 至 90°
    """

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None, pdf_path: Optional[str] = None,
                 match_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        :param pdf_path: (str) PDF 数据库的路径
        :param match_path: (str) 匹配规制的路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # PDF 数据库路径
        if pdf_path is not None:
            self.pdf_path = pdf_path
        else:
            self.pdf_path = os.path.join(XRD.Standard_Database, 'PDF_card')

        # PDF 卡片匹配规制路径
        if match_path is not None:
            self.match_path = match_path
        else:
            self.match_path = os.path.join(XRD.Standard_Database, 'PDF_card/PDF_match.xlsx')

        # self.__read_pdf
        self.pdf_data_dic = None
        self.pdf_name_all_list = None
        self.pdf_material_all_list = None
        self.pdf_cheformula_all_list = None

        # self.__match_xrd_pdf
        self.matching_rule = None

        self.matched_name_dic = None
        self.matched_material_dic = None
        self.matched_cheformula_dic = None

        self.xrd_data_list = None
        self.pdf_data_list = None
        self.xrd_name_list = None
        self.pdf_name_list = None

        self.pdf_material_list = None
        self.pdf_cheformula_list = None

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='XRD', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取 XRD 的源生 TXT 文件
        Read the raw TXT file of XRD.

        针对仪器： 德国布鲁克公司生产的X射线衍射仪

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        self.data.read_txt(txt_path=self.read_path, file_pattern=r"^(?!PDF#)")  # 读取 TXT 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, **kwargs) -> None:
        """
        绘制XRD 图像
        Plot the XRD image.

        :param save_path: (str) 文件的保存路径
        :param kwargs: 可加的关键词参数

        :return: None
        """

        self.data.plot_line(save_path=save_path, custom_yticks=[], **kwargs)  # 进行图形的绘制

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 绘制 XRD 和 PDF 的匹配图
    def plot_xrd_pdf(self, save_path: Union[bool, str] = True, width_height: tuple = (6, 4.5),
                     x_min: int = 5, x_max: int = 90, image_title: str = '', show_pdf: str = 'name',
                     xrd_color: Union[str, tuple] = '#B22222', pdf_color: Union[str, tuple] = '#000080',
                     xrd_background: str = 'Blues', pdf_background: str = 'Greys', print_detail: bool = True,
                     pdf_pattern: str = r'.*', dpi: int = 600) -> PrettyTable or None:
        """
        绘制 XRD 与 PDF 卡片图谱，需要先进行卡片的匹配
        Plot the XRD and PDF card spectra; card matching needs to be done first.

        :param save_path: (bool / str) 保存目录，若无赋值则用初始化中的 self.save_path，若为 False 或 None则为不保存
        :param width_height: (tuple) 图片的宽度和高度，默认为 (6, 4.5)
        :param x_min: (int) 横坐标的最小值
        :param x_max: (int) 横坐标的最大值
        :param image_title: (str) 图的标题，默认为无标题
        :param show_pdf: (str) 展示何种 PDF 卡片的内容，为 '' (空字符串) 时不展示，默认为 PDF 卡片的序列号
                         只能为 'name', 'material' or 'cheformula' 其中之一
        :param xrd_color: (str / tuple) 最上方线条的颜色，即 XRD 图像线条的颜色，默认为深红色
        :param pdf_color: (str / tuple) 下方线条的颜色，即 PDF 卡片的线条颜色，默认为深蓝色
        :param xrd_background: (str) 最上方背景的颜色，即 XRD 区域的背景色，默认为淡蓝色
        :param pdf_background: (str) 下方背景的颜色，即 PDF 卡片区域的背景色，默认为浅灰色
        :param print_detail: (bool) 是否打印 PDF 卡片的详细信息，默认为 True
        :param pdf_pattern: (str) 打印内容经过正则表达式筛选，默认为原内容
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600

        :return details: (PrettyTable) XRD 数据与 PDF 卡片配对的信息
        """

        # 当 save_path==True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            xrd_save_path = self.save_path
        # 若 save_path 为 None 或 False 时，本图形不保存
        elif save_path is False:
            xrd_save_path = None
        # 当有指定的 save_path 时，xrd_save_path 将会被其赋值，若 save_path == '' 则保存在运行的 py 文件的目录下
        else:
            xrd_save_path = save_path

        # 读取 PDF 卡片
        self.__read_pdf()
        # 匹配 PDF 卡片与 XRD 数据
        self.__match_xrd_pdf()

        # 检查 show_pdf 的赋值
        if show_pdf == 'name':
            target_list_list = self.pdf_name_list
        elif show_pdf == 'material':
            target_list_list = self.pdf_material_list
        elif show_pdf == 'cheformula':
            target_list_list = self.pdf_cheformula_list
        elif show_pdf == '':
            target_list_list = self.pdf_name_list  # 仅占位
        else:
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"show_pdf value must be one of 'name', 'material' or 'cheformula'.")

        # 绘制 XRD-PDF 图像 (此为上方 XRD 图像)
        for i, (xrd_data, pdf_data_list, target_list) in enumerate(
                zip(self.xrd_data_list, self.pdf_data_list, target_list_list)):

            # 设置子图的数量和高度比例
            num_subplots = len(pdf_data_list) + 1
            height_ratios = [4] + [1] * len(pdf_data_list)  # XRD 图形与 PDF 卡片的高度比例

            # 创建子图布局
            gs = GridSpec(nrows=num_subplots, ncols=1, height_ratios=height_ratios)
            # 创建图形布局并设置图片的宽度和高度
            fig = plt.figure(figsize=width_height, dpi=200)

            # 上方子图
            ax1 = fig.add_subplot(gs[0])  # 创建第一个子图，占据布局中的第一行
            ax1.plot(xrd_data['2θ / degree'], xrd_data['Intensity / (a.u.)'],  # 读取 2theta 列和 intensity 列的数据
                     color=xrd_color,  # 设置线条的颜色
                     linestyle='-')  # 设置折线条的类型
            ax1.set_xlabel('')  # 移除 X 轴标签

            if image_title != '':  # 设置标题
                plt.title(image_title, fontdict=self.data.font_title)

            # 设置 X 轴刻度
            plt.xlim((x_min, x_max))  # 限制横轴长度，默认只显示 5 至 90
            # 调整横轴的显示刻度
            start_range = math.ceil(x_min / 10) * 10
            end_range = math.floor(x_max / 10) * 10
            x_ticks = np.arange(start_range, end_range + 1, 10)  # 设置横轴的刻度，默认只显示 5 至 90

            plt.xticks(x_ticks, x_ticks)
            plt.setp(ax1.get_xticklabels(), visible=False)  # 移除X轴刻度标签
            ax1.set_yticks([])  # 移除 Y 轴刻度标签

            # 调用函数加背景，防止刻度被锁住
            general.Function.change_imshow(background_color=xrd_background, background_transparency=0.15,
                                        show_in_one=False)

            # 下方子图
            for j, (pdf_data, target) in enumerate(zip(pdf_data_list, target_list)):
                ax2 = fig.add_subplot(gs[j + 1], sharex=ax1)  # 创建后续的子图，共享 X 轴刻度
                # 读取 pdf_data 中的 2theta 列和 intensity 列的数据
                ax2.vlines(pdf_data['2θ / degree'], 0, pdf_data['Intensity / (a.u.)'],
                           colors=pdf_color)  # 设置线条的颜色
                ax2.set_ylim(0, 120)  # 设置 Y 轴范围
                ax2.set_yticks([])  # 移除 Y 轴刻度标签

                # 获取当前 X 轴的取值范围
                x_min, x_max = ax2.get_xlim()
                # 指定说明文字的 X 轴的位置
                target_x = x_max - 2

                # 在子图中添加说明文字
                if show_pdf:
                    try:
                        show_detail = re.findall(pdf_pattern, target)[0]  # 使用正则表达式查找匹配项
                    except IndexError:
                        show_detail = ''  # 如不满足正则表达式则不显示该内容
                    # 右对齐，上下居中
                    ax2.text(x=target_x, y=80, s=show_detail, fontdict=self.data.font_mark, ha='right', va='center')

                # 在最下方的子图上添加 X 轴标签
                if j == len(pdf_data_list) - 1:
                    ax2.set_xlabel('2θ / degree', fontdict=self.data.font_title)
                    # 设置 X 轴字体
                    plt.xticks(fontsize=18, fontweight='bold', fontfamily='Times New Roman')
                else:
                    plt.setp(ax2.get_xticklabels(), visible=False)  # 移除非最后一个子图的 X 轴刻度标签

                # 添加左侧 y_label
                fig.text(x=0.06, y=0.5, s='Intensity / (a.u.)', va='center', rotation='vertical',
                         fontdict=self.data.font_title)
                plt.subplots_adjust(bottom=0.15)

                # 调用函数加背景，防止刻度被锁住
                general.Function.change_imshow(background_color=pdf_background, background_transparency=0.15,
                                            show_in_one=False)

            # 调整布局
            plt.subplots_adjust(hspace=0)  # 移除子图之间的垂直间距

            # 保存图像
            title = self.xrd_name_list[i]
            if xrd_save_path is not None:  # 如果 xrd_save_path 的值不为 None，则保存
                file_name = title + ".png"  # 初始文件名为 "title.png"
                full_file_path = os.path.join(xrd_save_path, file_name)  # 创建完整的文件路径

                if os.path.exists(full_file_path):  # 查看该文件名是否存在
                    count = 1
                    file_name = title + f"_{count}.png"  # 若该文件名存在则在后面加 '_1'
                    full_file_path = os.path.join(xrd_save_path, file_name)  # 更新完整的文件路径

                    while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                        count += 1
                        file_name = title + f"_{count}.png"
                        full_file_path = os.path.join(xrd_save_path, file_name)  # 更新完整的文件路径

                plt.savefig(fname=full_file_path, dpi=dpi)  # 使用完整路径将散点图保存到指定的路径

            plt.show()  # 显示图形
            time.sleep(self.interval_time)  # 让程序休息一段时间，防止绘图过快导致程序崩溃

        # 打印 PDF 卡片的详细信息
        if print_detail:
            details = self.__print_pdf_details()

            return details

        else:
            return None

    # 绘制多个 XRD 对比图
    def plot_xrd_xrd(self, save_path: Union[bool, str] = True, width_height: Union[tuple] = (6, 4.5),
                     x_min: Union[int] = 5, x_max: Union[int] = 90, image_title: Union[str] = '',
                     line_color: Union[str, tuple, list, None] = None, add_line: bool = False,
                     dpi: Union[int] = 600) -> None:
        """
        绘制多个 XRD 对比图谱
        Plot multiple XRD comparative diffractograms

        :param save_path: (bool / str) 保存目录，若无赋值则用初始化中的 self.save_path，若为 False 或 None 则为不保存
        :param width_height: (tuple) 图片的宽度和高度，默认为 (6, 4.5)
        :param x_min: (int) 横坐标的最小值
        :param x_max: (int) 横坐标的最大值
        :param image_title: (str) 图的标题，默认为无标题
        :param line_color: (str / tuple / list) 最上方线条的颜色，即 XRD 图像线条的颜色，默认为红蓝渐变色
        :param add_line: (bool) 是否给每两个 XRD 曲线中添加横线
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600

        :return: None
        """

        # 当 save_path == True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            xrd_save_path = self.save_path
        # 若 save_path 为 None 或 False 时，本图形不保存
        elif save_path is False:
            xrd_save_path = None
        # 当有指定的 save_path 时，xrd_save_path 将会被其赋值，若 save_path == '' 则保存在运行的py文件的目录下
        else:
            xrd_save_path = save_path

        xrd_data_dic = copy.deepcopy(self.data.data_dic)

        # 计算图形布局所需的行数和列数
        num_rows = len(xrd_data_dic)
        num_cols = 1
        # 创建图形布局
        fig, axes = plt.subplots(num_rows, num_cols, figsize=width_height, sharex='all', dpi=200)

        # 设置线条颜色
        if isinstance(line_color, list):  # 输入 line_color 为 list 的情况
            # 其值需要大于或等于多 XRD 比较图数据数量，即 self.xrd_data_dic
            color_palette = line_color
        elif isinstance(line_color, str) or isinstance(line_color, tuple):  # 输入 line_color 为 str 的情况
            color_palette = [line_color for _ in xrd_data_dic]  # 列表中元素的数量与 self.xrd_data_dic 长度一致
        else:  # 无赋值的情况，或输入 line_color 为 None 的情况
            gradient_colors = [
                "#E60026",  # 红色
                "#CC0033",
                "#B30040",
                "#99004D",
                "#7F0066",
                "#660080",
                "#4C0099",
                "#3300B2"   # 蓝色
            ]

            color_palette = gradient_colors[:len(xrd_data_dic)]  # 根据名称列表的长度截取对应数量的颜色

        # 遍历数据并创建子图
        for i, (xrd_name, line_color) in enumerate(zip(xrd_data_dic, color_palette)):

            # 获取当前子图的坐标轴对象
            ax = axes[i]
            xrd_data = xrd_data_dic[xrd_name]
            # 绘制数据，读取 xrd_data 中的 2theta 列和 intensity 列的数据
            ax.plot(xrd_data['2θ / degree'], xrd_data['Intensity / (a.u.)'],
                    color=line_color,  # 设置线条的颜色
                    linestyle='-')  # 设置折线条的类型

            # 移除刻度标签
            ax.set_xticks([])
            ax.set_yticks([])
            # 去掉子图的标题
            ax.set_title('')

            # 移除中间的所有横线
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['bottom'].set_visible(False)
            ax.spines['left'].set_visible(False)

            # 添加外部框架
            if i == 0:  # 最上面一张图
                ax.spines['top'].set_visible(True)
                ax.spines['left'].set_visible(True)
                ax.spines['right'].set_visible(True)
                ax.set_title(image_title, fontdict=self.data.font_title)  # 添加标题

            elif i == num_rows - 1:  # 最下面一张图
                ax.spines['bottom'].set_visible(True)
                ax.spines['left'].set_visible(True)
                ax.spines['right'].set_visible(True)
                # 设置 X 轴刻度
                plt.xlim((x_min, x_max))  # 限制横轴长度，默认只显示5至90
                # 调整横轴的显示刻度
                start_range = math.ceil(x_min / 10) * 10
                end_range = math.floor(x_max / 10) * 10
                x_ticks = np.arange(start_range, end_range + 1, 10)  # 设置横轴的刻度，默认只显示5至90
                plt.xticks(x_ticks, x_ticks)
                # 设置 X 轴字体
                plt.xticks(fontsize=18, fontweight='bold', fontfamily='Times New Roman')

            else:  # 中间的图
                ax.spines['left'].set_visible(True)
                ax.spines['right'].set_visible(True)

            # 调整 Y 轴范围
            if i == 0:
                max_value = xrd_data['Intensity / (a.u.)'].max()
                ax.set_ylim(0, max_value * 1.2)
            elif i == num_rows - 1:
                max_value = xrd_data['Intensity / (a.u.)'].max()
                ax.set_ylim(max_value * -0.05, max_value * 1.05)
            else:
                max_value = xrd_data['Intensity / (a.u.)'].max()
                ax.set_ylim(0, max_value * 1.05)

            if add_line:
                # 在非最后一个子图上绘制横线
                if i < num_rows - 1:
                    ax.axhline(y=0, color='black', linestyle='-', linewidth=2)

        # 调整布局
        plt.subplots_adjust(hspace=0)  # 子图间距为 0

        # 添加整个图的 x_label 和 y_label
        fig.text(0.5, 0.02, '2θ / degree', ha='center', fontdict=self.data.font_title)
        fig.text(0.06, 0.5, 'Intensity / (a.u.)', va='center', rotation='vertical', fontdict=self.data.font_title)
        plt.subplots_adjust(bottom=0.15)  # 图片 (包括图的各种标签) 占用左和下部分的边距

        # 保存图像
        if xrd_save_path is not None:  # 如果 line_save_path 的值不为 None，则保存
            file_name = "XRD" + ".png"  # 初始文件名为 "title.png"
            full_file_path = os.path.join(xrd_save_path, file_name)  # 创建完整的文件路径

            if os.path.exists(full_file_path):  # 查看该文件名是否存在
                count = 1
                file_name = "XRD" + f"_{count}.png"  # 若该文件名存在则在后面加 '_1'
                full_file_path = os.path.join(xrd_save_path, file_name)  # 更新完整的文件路径

                while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                    count += 1
                    file_name = "XRD" + f"_{count}.png"
                    full_file_path = os.path.join(xrd_save_path, file_name)  # 更新完整的文件路径

            plt.savefig(fname=full_file_path, dpi=dpi)  # 使用完整路径将散点图保存到指定的路径

        plt.show()  # 显示图形
        time.sleep(self.interval_time)  # 让程序休息一段时间，防止绘图过快导致程序崩溃

        return None

    # 添加 XRD 与 PDF 匹配
    def add_xrd_pdf_match(self, directory: Optional[str] = None, match_path: Optional[str] = None,
                          save_pdf: bool = True) -> None:
        """
        添加新的 XRD 与 PDF 配对至 Excel中

        :param directory: (str) XRD 文件与 PDF 卡片文件的文件夹的路径
        :param match_path: (str) 匹配 Excel 的路径，默认为 Standard_Database 路径
        :param save_pdf: (bool) 是否复制 PDF 卡片至 Excel 的目录下，只有成功添加至 Excel 匹配表格后才会查看是否复制卡片，默认为 True

        :return: None
        """

        # 检查 directory 是否为 None 或者不是一个有效的目录
        if directory is None:
            directory = self.read_path
        elif not os.path.isdir(directory):
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"The provided directory is not valid or is None.")

        # PDF 卡片匹配规制路径
        if match_path is None:
            match_path = self.match_path

        # 检索目录中的所有文件
        all_files = os.listdir(directory)

        # 筛选出以 'PDF#' 开头和非 'PDF#' 开头的 txt 文件，并去除扩展名
        pdf_files = [file.split('.')[0] for file in all_files if file.startswith('PDF#') and file.endswith('.txt')]
        non_pdf_files = [file.split('.')[0] for file in all_files if
                         not file.startswith('PDF#') and (file.endswith('.txt') or file.endswith('.TXT'))]

        # 检查文件数量是否符合要求
        if len(non_pdf_files) != 1:
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"There should be exactly one non-'PDF#' txt file.")
        if len(pdf_files) == 0:
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"There should be at least one 'PDF#' txt file.")

        # 读取 Excel 文件
        df = pd.read_excel(match_path)

        # 检查第一列是否已经有指定的非 'PDF#' 文件
        if non_pdf_files[0] in df.iloc[:, 0].values:
            print(f"The txt file \033[33m{non_pdf_files[0]}\033[0m already exists in the Excel file.")

            return None

        # 准备新的行数据
        new_row = [non_pdf_files[0]] + pdf_files

        # 检查并扩展列数
        required_columns = len(new_row)
        current_columns = df.shape[1]
        if required_columns > current_columns:
            # 如果需要的列数比现有的列数多，则添加缺失的列
            for i in range(current_columns, required_columns):
                df[f'Column_{i}'] = None

        # 添加新的行到 DataFrame
        new_row_df = pd.DataFrame(data=[new_row], columns=df.columns[:required_columns])
        df = pd.concat(objs=[df, new_row_df], ignore_index=True)

        # 保存并关闭 Excel 文件
        df.to_excel(match_path, index=False)

        # 使用 openpyxl 调整列宽和填充颜色
        workbook = load_workbook(match_path)
        worksheet = workbook.active

        # 设置第一列的宽度为 27
        worksheet.column_dimensions[get_column_letter(1)].width = 27

        # 设置其他列的宽度为 17
        for col in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 17

        # 设置填充颜色
        light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        light_green_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

        # 第一列设置为淡蓝色
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = light_blue_fill

        # 其他列设置为淡绿色
        for col in range(2, worksheet.max_column + 1):
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.fill = light_green_fill

        # 保存调整后的工作簿
        workbook.save(match_path)

        # 打印完成信息，包括非 'PDF#' txt 文件的名称
        print(f"The txt file \033[33m{non_pdf_files[0]}\033[0m has been added to the Excel file.")

        # 复制 PDF 卡片到 PDF 卡片数据库中
        if save_pdf:

            pdf_dir = os.path.dirname(match_path)  # 匹配的 Excel 所在的目录
            for pdf_file in pdf_files:

                pdf_original_path = os.path.join(directory, pdf_file) + ".txt"  # 旧路径
                pdf_target_path = os.path.join(pdf_dir, pdf_file) + ".txt"  # 新路径
                # 检查目标文件是否存在
                if os.path.exists(pdf_target_path):
                    print(f"The \033[92m{pdf_file}\033[0m already exists")
                else:
                    shutil.copyfile(pdf_original_path, pdf_target_path)
                    print(f"\033[92m{pdf_file}\033[0m copy succeeded")

        return None

    # 检查 PDF 卡片
    def inspect_pdf(self, pdf_inspect_path: Optional[str] = None, print_detail: bool = True,) -> PrettyTable:
        """
        检查目标目录及所有子目录下所有的 PDF 卡片，并查看每个 PDF 卡片出现的次数
        Check all PDF cards in the target directory and all subdirectories,
        and see how many times each PDF card appears

        :param pdf_inspect_path: (str) PDF 需要检查的 PDF 卡片目录的路径，默认为 Standard_Database 数据库的路径
        :param print_detail: (bool) 是否打印 PDF 卡片详细信息，默认为 True

        :return table: (PrettyTable) PDF 数据库中所有卡片的信息
        """

        if pdf_inspect_path is None:
            pdf_inspect_path = self.pdf_path

        # 遍历目录及其子目录，寻找以 'PDF#' 开头并以 '.txt' 结尾的文件
        txt_files = [file.stem for file in Path(pdf_inspect_path).rglob('PDF#*.txt')]
        txt_paths = [file for file in Path(pdf_inspect_path).rglob('PDF#*.txt')]

        # 计算每个文件名的出现次数
        file_counts = Counter(txt_files)

        pdf_material_all_dict = {}  # 用字典存储每个文件的材料名和化学式
        for txt_path in txt_paths:
            # 读取 TXT 文件
            with open(txt_path, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()

                pdf_material = lines[1].strip()  # 获取 PDF 卡片的材料名 (在 TXT 文件的第二行)
                pdf_cheformula = lines[2].strip()  # 获取 PDF 卡片的化学式 (在 TXT 文件的第三行)

                # 将材料名和化学式添加到字典
                pdf_material_all_dict[txt_path.stem] = (pdf_material, pdf_cheformula)

        # 按计数降序排序
        sorted_file_counts = sorted(file_counts.items(), key=lambda x: x[1], reverse=True)

        # 创建 PrettyTable，包括文件名、计数、材料名和化学式
        table = PrettyTable(["File Name", "Count", "Material Name", "Chemical Formula"])
        for file_name, count in sorted_file_counts:
            material, cheformula = pdf_material_all_dict.get(file_name, ("N/A", "N/A"))
            table.add_row([file_name, count, material, cheformula])

        # 打印 PDF 卡片详细信息
        if print_detail:
            print(table)

        return table

    # 读取 PDF 卡片
    def __read_pdf(self) -> Tuple[Dict[str, DataFrame], List[str], List[str], List[str]]:
        r"""
        读取PDF卡片的数据，打开的 self.pdf_path 为 PDF 卡片所在目录路径
        Reading data from the PDF card, where self.pdf_path is the directory path to the PDF card.

        对于全路径：r'\/.*\/(?P<pdf_name>PDF#\d+-\d+-\d+)\.txt$'
        对于相对路径：os.txt_path.splitext(os.txt_path.basename(file_path))[0]

        :return dictionary：(dict) 含有该目录下所有 TXT 文件数据的 key & value，key 为 PDF 卡片的序号，value 为 DataFrame 表格
        :return pdf_name_all_list：(list) PDF 卡片的序列号
        :return pdf_material_all_list：(list) PDF 卡片的材料名
        :return pdf_Cheformula_list：(list) PDF 卡片的化学式
        DataFrame: 2theta:       float64
                   intensity:    float64
                   dtype:        object
        """

        pdf_data_dic = {}  # 创造一个空的 dict 以存储 PDF 卡片的数据
        pdf_name_all_list = []  # 创造一个空的 list 以存储 PDF 卡片的序号
        pdf_material_all_list = []  # 创造一个空的 list 以存储 PDF 卡片的材料名
        pdf_cheformula_all_list = []  # 创造一个空的 list 以存储 PDF 卡片的化学式

        # 遍历目录下所有 TXT 文件，并将它们读取为 DataFrame 并存入列表中
        for file_path in os.listdir(self.pdf_path):

            if file_path.startswith('PDF#') and file_path.endswith('.txt'):  # 筛选以 'PDF#' 开头且以 'txt' 结尾的文件

                pdf_txt_path = os.path.join(self.pdf_path, file_path)

                # 使用 splitext() 分割文件名和扩展名
                pdf_name_all, _ = os.path.splitext(file_path)
                pdf_name_all_list.append(pdf_name_all)

                # 读取 TXT 文件
                with open(pdf_txt_path, 'r', encoding='iso-8859-1') as file:
                    lines = file.readlines()

                    cleaned_lines = []
                    for line in lines:
                        cleaned_line = line.lstrip(' ')  # 删除每行开头的空格
                        cleaned_lines.append(cleaned_line)

                    pdf_material = lines[1].strip()  # 获取 PDF 卡片的材料名 (在 TXT 文件的第二行)
                    pdf_cheformula = lines[2].strip()  # 获取 PDF 卡片的化学式 (在 TXT 文件的第三行)

                # 找到有效行
                pattern = re.compile(r'^\d+\.\d\d\d')
                valid_lines = [line for line in cleaned_lines if pattern.match(line)]
                # 从有效行创建 DataFrame
                pdf_data_df = pd.DataFrame([line.strip().split() for line in valid_lines])
                # 获取第一列 (2θ) 和第三列的数据 (I(f))，并将 '<1' 替换成 0.5
                pdf_data_df = pdf_data_df.iloc[:, [0, 2]].replace('<1', 0.5)
                # 更改列索引的名称为：2theta  intensity
                pdf_data_df = pdf_data_df.rename(columns={0: '2θ / degree', 2: 'Intensity / (a.u.)'})
                # 将其中的数据类型从 object 转换为 float64
                pdf_data_df = pdf_data_df.astype(float)
                # 将 PDF 卡片的数据传入到 pdf_data_dic 中key 为 PDF 卡片的序号，value 为 DataFrame 表格
                pdf_data_dic[os.path.splitext(os.path.basename(file_path))[0]] = pdf_data_df
                # 将 PDF 卡片的材料名传入到 pdf_material_all_list 中
                pdf_material_all_list.append(pdf_material)
                # 将化学式加入到 pdf_cheformula_all_list 中
                pdf_cheformula_all_list.append(pdf_cheformula)

        self.pdf_data_dic = pdf_data_dic  # PDF 卡片数据的 dict 传给 self.pdf_data_dic
        self.pdf_name_all_list = pdf_name_all_list  # PDF 卡片序列号的 list 传给 self.pdf_name_all_list
        self.pdf_material_all_list = pdf_material_all_list  # PDF 卡片材料名的 list 传给 self.pdf_material_all_list
        self.pdf_cheformula_all_list = pdf_cheformula_all_list  # PDF 卡片的化学式 list 传给 self.pdf_cheformula_all_list

        return pdf_data_dic, pdf_name_all_list, pdf_material_all_list, pdf_cheformula_all_list

    # 匹配 XRD 图谱 与 PDF 卡片
    def __match_xrd_pdf(self) -> Tuple[Dict[str, List[str]],
                               List[str], List[List[DataFrame]], List[str], List[List[str]]]:
        """
        打开匹配规则的表格，并使得 PDF 卡片与 XRD 数据相匹配
        Open the table with matching rules and align the PDF card with the XRD data.

        分为两个部分:1.  建立 XRD 与 PDF 对应库：这个库有英文和拼音两种解析形式
                   2.  XRD 名称与 PDF 卡片的匹配：从 self.xrd_data_dic 为起点，提取 xrd_name_base，查看是否在 matching_rule 中，
                       从相应位置提取
                   注意: 匹配时只能匹配小写字母 [a-z], 如果匹配有匹配项而没有导入 PDF 卡片时会报错

        :return matching_rule： (dict) dict，PDF 卡片与 XRD 数据的匹配规则。例：{'pdf_name1': ['PDF#number1', 'PDF#number2]}
        :return matched_dic： (dict) dict，PDF 卡片与 XRD 已配对的数据。例：{'pdf_name1': ['PDF#number1', 'PDF#number2]}
        :return xrd_data_list： (list) list，用以存储 DataFrame 形式的 xrd_data。例：[xrd_df1, xrd_df2]
        :return pdf_data_list： (list) list，用以存储 PDF 卡片的 DataFrame 组成的 list。例：[[pdf_df1, pdf_df2], [pdf_df3]]
        :return xrd_name_list： (list) list，用以存储 XRD 文件的名称。例：[xrd_name1, xrd_name2]
        :return pdf_name_list： (list) list， 用以存储 PDF 卡片文件的名称。例：[[pdf_name1, pdf_name2], [pdf_name3]]
                xrd_data_list，pdf_data_list，xrd_name_list，pdf_name_list 此四个 list 内数据长度要一致，且顺序对应
        """

        # 第一部分：创建空的 dict 和 list 用以存储数据
        matching_rule = {}

        matched_name_dic = {}  # 创建空的 dict，用以存储配对后的 XRD-PDF 数据
        matched_material_dic = {}  # 创建空的 dict，用以存储配对后的材料名数据
        matched_cheformula_dic = {}  # 创建空的 dict，用以存储配对后的化学式数据

        xrd_data_list = []  # 创建空的 list，用以存储 DataFrame 形式的 xrd_data
        pdf_data_list = []  # 创建空的 list，用以存储 PDF 卡片的 DataFrame 组成的 list
        xrd_name_list = []  # 创建空的 list，用以存储 XRD 文件的名称
        pdf_name_list = []  # 创建空的 list，用心存储 PDF 卡片名称

        pdf_material_list = []  # 创建空的 list，用以存储 XRD 文件的材料名
        pdf_cheformula_list = []  # 创建空的 list，用以存储 XRD 文件的化学式

        # 第二部分：打开 PDF 卡片与 XRD 数据对应规则的库
        xrd_pdf_df = pd.read_excel(self.match_path, sheet_name=0, header=0, index_col=None, nrows=None)
        for index, row in xrd_pdf_df.iterrows():
            if isinstance(row['material'], str):  # 材料名的那一列
                material = row['material'].strip()
            else:
                material = row['material']
            pdfs = [pdf.strip() for pdf in row[1:] if pd.notnull(pdf)]
            matching_rule[material] = pdfs

        xrd_data_dic = copy.deepcopy(self.data.data_dic)

        # 第三部分：XRD 名称与 PDF 卡片的匹配
        for xrd_name in xrd_data_dic.keys():

            if xrd_name in matching_rule.keys():
                matched_name_dic[xrd_name] = matching_rule[xrd_name]  # 赋值 matched_dic
                xrd_data_list.append(xrd_data_dic[xrd_name])  # 将 xrd_name_all 对应的 DataFrame 存入 xrd_name_list
                pdf_name_s = matching_rule[xrd_name]
                pdf_data = [self.pdf_data_dic[pdf_name] for pdf_name in pdf_name_s]
                pdf_data_list.append(pdf_data)  # 将 xpdf_data 存入 pdf_data_list
                xrd_name_list.append(xrd_name)  # 将 xrd_name_all 存入 xrd_name_list
                pdf_name_list.append(pdf_name_s)  # 将 xpdf_data 对应的 list_DataFrame 存入 pdf_data_list

        # 第四部分：传递配对的 XRD-PDF 数据
        for key, values in matched_name_dic.items():
            # 初始化新的列表
            matched_materials = []
            matched_cheformulas = []

            # 遍历每一个 PDF 编号
            for value in values:

                # 如果这个 PDF 编号在 pdf_name_all_list 中，找到对应的位置
                if value in self.pdf_name_all_list:

                    index = self.pdf_name_all_list.index(value)
                    # 使用相同的位置获取 pdf_material_list 和 pdf_cheformula_all_list 中的元素
                    matched_materials.append(self.pdf_material_all_list[index])
                    matched_cheformulas.append(self.pdf_cheformula_all_list[index])

            # 将结果放入新的 dict 和 list 中
            matched_material_dic[key] = matched_materials
            matched_cheformula_dic[key] = matched_cheformulas

            pdf_material_list.append(matched_materials)
            pdf_cheformula_list.append(matched_cheformulas)

        # 第五部分将数据赋值给实例属性
        self.matching_rule = matching_rule

        self.matched_name_dic = matched_name_dic
        self.matched_material_dic = matched_material_dic
        self.matched_cheformula_dic = matched_cheformula_dic

        self.xrd_data_list = xrd_data_list
        self.pdf_data_list = pdf_data_list
        self.xrd_name_list = xrd_name_list
        self.pdf_name_list = pdf_name_list

        self.pdf_material_list = pdf_material_list
        self.pdf_cheformula_list = pdf_cheformula_list

        return matching_rule, xrd_data_list, pdf_data_list, xrd_name_list, pdf_name_list

    # 打印 PDF 卡片信息
    def __print_pdf_details(self) -> PrettyTable:
        """
        打印 XRD 与 PDF 卡片对应的详细信息
        Print XRD and PDF card corresponding details

        :return table: (PrettyTable) XRD 与 PDF 卡片对应的详细信息
        """

        # 创建 PrettyTable 表格
        table = PrettyTable()
        table.field_names = ["Sample", "PDF_card", "Material", "Chemical Formula"]

        # 假设所有字典都有相同的键，并且对应的值的列表长度也相同
        for key in self.matched_name_dic:
            # 获取每个字典中键对应的值的长度
            num_entries = len(self.matched_name_dic[key])
            # 计算样品名称应该出现在哪一行（中间位置）
            sample_row = num_entries // 2
            for i in range(num_entries):
                # 从每个字典中获取相应的值
                name = self.matched_name_dic[key][i]
                material = self.matched_material_dic[key][i]
                cheformula = self.matched_cheformula_dic[key][i]

                # 将样品名称放在中间位置
                row_sample = key if i == sample_row else ""
                table.add_row([row_sample, name, material, cheformula])

            # 在每个样品之后添加分隔行，除了最后一个样品
            if key != list(self.matched_name_dic.keys())[-1]:
                table.add_row(["-" * 30, "-" * 20, "-" * 20, "-" * 20])  # 使用适当数量的破折号以适应列宽

        # 打印表格
        print(table)

        return table


""" Raman """
class Raman(Keyword):
    """
    默认参数：

    测试区间为波长 100 至 2000
    """

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 拟合曲线
        self.smoothing_dic = None
        self.key_point_dic = None

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='Raman', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取 Raman 的源生 TXT 文件
        Read the raw TXT file of Raman.

        针对仪器： 英国雷尼绍公司生产的显微共聚焦激光拉曼光谱仪

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        self.data.read_txt(txt_path=self.read_path)  # 读取 TXT 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, add_fitting: bool = False, degree: int = 3,
             smoothing: float = 0.05, alpha: float = 0.8, show_peak: bool = False, **kwargs) -> str:
        """
        绘制 Raman 的图像
        Plot the Raman image.

        :param save_path: (str) 文件的保存路径
        :param add_fitting: (bool) 是否添加拟合曲线，默认 False
        :param degree: (int) 曲线阶数，越大越平滑，1 <= degree_fitting <= 5
        :param smoothing: (float) 平滑因子，越小越接近原数据
        :param alpha: (float) 线条的透明程度，越小越透明，默认为 0.8
        :param show_peak: (bool) 是打印峰值点，只有 add_fitting==True 时才有意义，默认为 False
        :param kwargs: 可加的关键词参数

        :return details: (str) Raman 数据的关键点信息
        """

        details = ''  # 构建空字符串

        # 判断是否添加光滑曲线
        if add_fitting:

            # 调用私有方法 __curve_fitting() 来获取光滑的 DataFrame 数据和关键点
            smoothing_dic, key_point_dic = self.__curve_fitting(degree, smoothing)

            # 使用 zip 循环遍历三个 dict 的 items
            for (k1, v1), (k2, v2), (k3, v3) in \
                    zip(self.data.data_dic.items(), smoothing_dic.items(), key_point_dic.items()):

                temp_dict = {k1: v1, k2: v2}

                # 绘制图像
                self.data.plot_line(data_dic=temp_dict,  save_path=save_path,
                                    alpha=alpha,  show_in_one=True,
                                    custom_yticks=[], **kwargs)  # 进行图形的绘制

                # 判断是否打印关键点
                if show_peak:

                    # 将包含 DataFrame 的字典转换为可以被 json.dumps 处理的格式
                    v3_serializable = {key: value.to_dict() if isinstance(value, pd.DataFrame)
                    else value for key, value in v3.items()}

                    # 现在 v3_serializable 可以被 json.dumps 处理
                    v3_string = json.dumps(v3_serializable, indent=4)

                    # 构建 detail 字符串
                    detail = '-' * 50 + f'\033[91m{k3}\033[0m' + '-' * 50 + '\n' + v3_string + '\n'

                    print(detail)  # 打印单个 detail

                    details += detail  # 构建全体 details

            return details

        # 不添加光滑曲线的情况
        else:

            # 绘制图像
            self.data.plot_line(save_path=save_path, custom_yticks=[], **kwargs)  # 进行图形的绘制

            return details

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 进行曲线拟合
    def __curve_fitting(self, degree: int = 3, smoothing: float = 0.05)\
            -> Tuple[Dict[str, DataFrame], Dict[str, DataFrame]]:
        """
        进行曲线拟合，如果 add_fitting == True 则调用此函数
        Perform curve fitting; call this function if add_fitting == True.

        :param degree: (int) 曲线阶数，越大越平滑，1 <= degree_fitting <= 5
        :param smoothing: (float) 平滑因子，越小越接近原数据

        :return smoothing_dic: (dict) 保存光滑后曲线的 dict
        :return key_point_dic: (dict) 保存关键点的 dict
        """

        data_dic = copy.deepcopy(self.data.data_dic)

        # 实例化 general.Magic 来光滑曲线
        self.curve_smoothing = general.Magic(keyword='Raman', data_dic=data_dic)
        self.curve_smoothing.smooth_curve(degree_smoothing=degree,  smooth_smoothing=smoothing)
        self.curve_smoothing.improve_precision(spline_dic=self.curve_smoothing.spline_smoothing_dic,
                                               precision_smoothing=1)

        # 实例化 general.Magic 来找出关键点
        self.curve_point = general.Magic(data_dic=self.curve_smoothing.data_dic)
        self.curve_point.smooth_curve(degree_smoothing=degree, smooth_smoothing=smoothing)
        self.curve_point.locate_point(precision_dxdy=0.1, locate_inflection=False,
                                      locate_extremum=True, locate_max=True, locate_min=False)

        smoothing_dic = {k + "_fitting": v for k, v in copy.deepcopy(self.curve_smoothing.smoothing_dic).items()}
        key_point_dic = copy.deepcopy(self.curve_point.key_point_dic)

        self.smoothing_dic = smoothing_dic
        self.key_point_dic = key_point_dic

        return smoothing_dic, key_point_dic


""" DIL """
class DIL(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='DIL', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取热膨胀分析的源生 TXT 文件
        Read the raw TXT file of the thermal expansion analysis.

        针对仪器： 德国耐驰仪器制造有限公司生产的热膨胀仪

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   dtype:         object
        """

        self.data.read_txt(txt_path=self.read_path)  # 读取 TXT 文件数据
        data_dic = copy.deepcopy(self.data.data_dic)

        # 对 data_dic 中每个 DataFrame 的第二列乘以1000
        for key, df in data_dic.items():
            df.iloc[:, 1] = df.iloc[:, 1] * 1000

        self.data.data_dic = data_dic

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, show_in_one: bool = False, show_peak: bool = True, **kwargs)\
            -> None:
        """
        绘制热膨胀分析的图像
        Plot the image of thermal expansion analysis.

        :param save_path: (str) 文件的保存路径
        :param show_in_one: (bool) 是否将所有图绘制在一张图像中，默认为否
        :param show_peak: (bool) 是否打印峰值信息，默认为 True
        :param kwargs: 可加的关键词参数

        :return: None
        """

        # 更改 Y 轴标题
        y_label = r'dL/L$_{0}$ × 10$^{-3}$'
        self.data.plot_line(save_path=save_path, y_label=y_label, show_in_one=show_in_one, **kwargs)  # 进行图形的绘制

        if show_peak:
            for title, data_df in self.data.data_dic.items():

                peaks, _ = find_peaks(data_df.iloc[:, 1].values)  # 使用 find_peaks 找到峰值

                # 获取最大峰值对应的温度
                peak_temperature = data_df.iloc[:, 0].iloc[peaks[np.argmax(data_df.iloc[:, 1].iloc[peaks])]]

                # 获取最大峰值对应的 dL_L0 值
                peak_value = data_df.iloc[:, 1].iloc[peaks[np.argmax(data_df.iloc[:, 1].iloc[peaks])]]

                print(f"In \033[92m{title:<20}\033[0m, "
                      f"The peak occurs at \033[94m{peak_temperature:7.2f}°C\033[0m "
                      f"with a value of \033[95m{peak_value:4.2f}dL/Lo\033[0m")

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 绘制带一阶导数的曲线
    def plot_derivative(self, save_path: Union[bool, str] = True, show_peak: bool = True,
                        left_color: Union[str, tuple] = '#FF6347', left_style: str = '-',
                        right_color: Union[str, tuple] = '#6495ED', right_style: str = '-.',
                        background_color: Optional[str] = 'Greens', background_transparency: Optional[float] = 0.15,
                        dpi: int = 600, **kwargs) -> None:
        """
        绘制热膨胀及其一阶导数曲线
        Plotting the Thermal Expansion and Its First Derivative Curve.

        :param save_path: (str) 文件的保存路径
        :param show_peak: (bool) 是否打印峰值信息，默认为 True
        :param left_color: (str / tuple) 左轴曲线，即原始曲线的颜色，默认为亮红色 (Tomato色)
        :param left_style: (str) 左轴曲线，即原始曲线的风格，默认为线条
        :param right_color: (str / tuple) 右轴曲线，即一阶导数曲线的颜色，默认为亮蓝色 (Cornflower Blue色)
        :param right_style: (str) 右轴曲线，即一阶导数曲线的风格，默认为点画线
        :param background_color: (str / tuple) 设置图片的背景颜色，默认为 'Greens'
        :param background_transparency: (float) 背景色的透明度，只有存在背景色时才有意义，默认为 0.15
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600

        :return: None

        --- **kwargs ---

        - x_min: (float) X 轴的最小值，默认为 None
        - x_max: (float) X 轴的最大值，默认为 None
        - left_min: (float) 左轴的最小值，默认为 None
        - left_max: (float) 左轴的最大值，默认为 None
        - right_min: (float) 右轴的最小值，默认为 None
        - right_max: (float) 右轴的最大值，默认为 None
        """

        # 检查赋值
        if True:

            data_dic = copy.deepcopy(self.data.data_dic)
    
            # 当 save_path == True 时，沿用 self.save_path 的设置，此项为默认项
            if save_path is True:
                save_path = self.save_path
            # 若 save_path 为 False 时，本图形不保存
            elif save_path is False:
                save_path = None
            # 当有指定的 save_path 时，line_save_path 将会被其赋值，若 save_path == '' 则保存在运行的py文件的目录下
            else:
                save_path = save_path

            # 关键字参数初始化
            x_min = kwargs.pop('x_min', None)
            x_max = kwargs.pop('x_max', None)
            left_min = kwargs.pop('left_min', None)
            left_max = kwargs.pop('left_max', None)
            right_min = kwargs.pop('right_min', None)
            right_max = kwargs.pop('right_max', None)

        derivatives_dic = {}
        for title, df in data_dic.items():

            # 使用iloc获取温度和dL/L0的数据，并计算一阶导数
            temperatures = df.iloc[:, 0]  # 假设温度是第一列
            values = df.iloc[:, 1]  # 假设dL/L0是第二列
            derivatives = np.gradient(values, temperatures)  # 一阶导数

            data = {'Temperature/°C': temperatures, 'derivatives': derivatives}
            df_new = pd.DataFrame(data)
            derivatives_dic[title] = df_new

            x_label = r'Temperature/°C'
            y1_label = r'dL/L$_{0}$ × 10$^{-3}$'
            y2_label = r'dL/dt × 10$^{-3}$'

            # 创建图形
            fig = plt.figure(figsize=(7, 4.5), dpi=200)

            # 添加第一个子图
            ax1 = fig.add_subplot(111)
            ax1.set_xlabel(x_label, fontdict=self.data.font_title)
            # 调整 X 轴刻度字体、大小并加粗
            for label in ax1.get_xticklabels():
                label.set_fontproperties(self.data.font_title)

            # 将左轴用于绘制原始数据
            ax1.plot(temperatures, values,
                     color=left_color,
                     linestyle=left_style)
            ax1.set_ylabel(y1_label, color=left_color, fontdict=self.data.font_title)
            ax1.tick_params(axis='y', labelcolor=left_color, labelsize=self.data.font_ticket['size'])
            # 调整 Y 左轴刻度的字体、大小并加粗
            plt.yticks(fontsize=self.data.font_ticket['size'],
                       fontweight=self.data.font_ticket['weight'],
                       fontfamily=self.data.font_ticket['family'])

            # 添加第二个 Y 轴用于一阶导数
            ax2 = ax1.twinx()
            ax2.plot(temperatures, derivatives,
                     color=right_color,
                     linestyle=right_style)
            ax2.set_ylabel(y2_label, color=right_color, fontdict=self.data.font_title)
            ax2.tick_params(axis='y', labelcolor=right_color, labelsize=self.data.font_ticket['size'])
            # 调整 Y 右轴刻度的字体、大小并加粗
            plt.yticks(fontsize=self.data.font_ticket['size'],
                       fontweight=self.data.font_ticket['weight'],
                       fontfamily=self.data.font_ticket['family'])

            # 设置 X 轴
            plt.xlim((x_min, x_max))

            # 设置左轴的最小值和最大值
            ax1.set_ylim(bottom=left_min, top=left_max)

            # 设置右轴的最小值和最大值
            ax2.set_ylim(bottom=right_min, top=right_max)

            general.Function.change_imshow(background_color, background_transparency)  # 添加背景

            plt.tight_layout()  # 调整布局

            # 如果提供了保存路径，则保存图像到指定路径
            if save_path is not None:  # 如果 save_path 的值不为 None，则保存
                file_name = title + ".png"  # 初始文件名为 "title.png"
                full_file_path = os.path.join(save_path, file_name)  # 创建完整的文件路径

                if os.path.exists(full_file_path):  # 查看该文件名是否存在
                    count = 1
                    file_name = title + f"_{count}.png"  # 若该文件名存在则在后面加 '_1'
                    full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                    while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                        count += 1
                        file_name = title + f"_{count}.png"
                        full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                plt.savefig(fname=full_file_path, dpi=dpi)  # 使用完整路径将散点图保存到指定的路径

            plt.show()  # 显示图形
            time.sleep(self.interval_time)  # 让程序休息一段时间，防止绘图过快导致程序崩溃

        if show_peak:
            for title, data_df in data_dic.items():

                peaks, _ = find_peaks(data_df.iloc[:, 1].values)  # 使用 find_peaks 找到峰值

                # 获取最大峰值对应的温度
                peak_temperature = data_df.iloc[:, 0].iloc[peaks[np.argmax(data_df.iloc[:, 1].iloc[peaks])]]

                # 获取最大峰值对应的 dL_L0 值
                peak_value = data_df.iloc[:, 1].iloc[peaks[np.argmax(data_df.iloc[:, 1].iloc[peaks])]]

                print(f"In \033[92m{title:<20}\033[0m, "
                      f"The peak occurs at \033[94m{peak_temperature:7.2f}°C\033[0m "
                      f"with a value of \033[95m{peak_value:4.2f}dL/Lo\033[0m")

                derivatives_df = derivatives_dic[title]

                peaks, _ = find_peaks(derivatives_df.iloc[:, 1].values)  # 使用 find_peaks 找到峰值

                # 获取最大峰值对应的温度
                peak_temperature = \
                    derivatives_df.iloc[:, 0].iloc[peaks[np.argmax(derivatives_df.iloc[:, 1].iloc[peaks])]]

                # 获取最大峰值对应的 dL_L0 值
                peak_value = derivatives_df.iloc[:, 1].iloc[peaks[np.argmax(derivatives_df.iloc[:, 1].iloc[peaks])]]

                print(f"In \033[92m{title:<20}\033[0m, "
                      f"The \033[33;2mderivative peak\033[0m occurs at \033[94m{peak_temperature:7.2f}°C\033[0m "
                      f"with a value of \033[95m{peak_value:4.2f}dL/Lo\033[0m")


""" XPS """
class XPS(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # 数据分离
        self.data_valid_dic = None
        self.data_black_dic = None

        self.y_result_df = None

        # 拟合
        self.smoothing_dic = None
        self.smoothing_original_precision_dic = None
        self.realized_dic = None
        self.realized_original_precision_dic = None

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='XPS', save_path=save_path)

    # 读取原生文件
    def read(self) -> Dict[str, DataFrame]:
        """
        读取 XPS 的源生 Excel 文件，如果文件的扩展名为 'xls'，则先改为 'xlsx'
        Read the raw EXCEL file of the XPS; if the file extension is 'xls', change it to 'xlsx' first.

        针对仪器： 采用中国赛默飞世尔科技有限公司生产的 X 射线光电子能谱议

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   column3:       float64
                   dtype:         object
        """

        # 更改 Excel 的扩展名，若无需要更改的也不影响，只有为路径为目录的情况下才会更改
        if os.path.isdir(self.read_path):
            self.data.rename_extension(path=self.read_path, old_extension='xls', new_extension='xlsx', show=False)

        self.data.read_excel(excel_path=self.read_path)  # 读取 Excel 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        data_valid_dic = {}
        data_black_dic = {}

        # 遍历原 dict 中的每个键值对
        for title, df in data_dic.items():

            # 提取 valid data & black data，并分别保存到新的 dict 中
            data_valid_dic[title] = df.iloc[:, [0, 1]]
            data_black_dic[title] = df.iloc[:, [0, 2]]

        self.data_valid_dic = data_valid_dic
        self.data_black_dic = data_black_dic

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, fitting: bool = False, noise_level: float = 0,
             degree: int = 3, smoothing: float = 200, deconvolution: Union[tuple, list, None] = None,
             area: Union[tuple, list, None] = None, high: Union[tuple, list, None] = None, precision: float = 0.2,
             show_perfect: bool = False, show_legend: bool = False, show_information: bool = True, dpi: int = 600,
             **kwargs) -> None:
        """
        绘制 XPS 的图像
        Plot the XPS image.

        :param save_path: (str) 文件的保存路径
        :param fitting: (bool) 是否添加拟合曲线，默认 False
        :param noise_level: (float) 添加的噪声程度，只有曲线进行拟合此项才有意义，此项必需大于等于 0，默认为 0，即不添加
        :param degree: (int) 曲线阶数，越大越平滑，1 <= degree_fitting <= 5
        :param smoothing: (float) 平滑因子，越小越接近原数据
        :param deconvolution: (tuple / list) 选择出峰的位置，为 None 时不进行反卷积
        :param area: (tuple / list) 反卷积峰的面积，长度需要与 deconvolution 相同
        :param high: (tuple / list) 反卷积峰的高度，长度需要与 deconvolution 相同
        :param precision: (float) 绘制散点的精度，默认为 0.02
        :param show_perfect: (bool) 让散点贴近拟合出的曲线，此项可以加噪声，默认为 False
        :param show_legend: (bool) 显示散点 / 曲线信息，默认为 False
        :param show_information: (bool) 打印峰值信息，默认为 True
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600
        :param kwargs: 添加峰的曲线关键字参数

        :return: None
        """

        # 当 save_path==True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            save_path = self.save_path
        # 若 save_path 为 False 时，本图形不保存
        elif save_path is False:
            save_path = None
        # 当有指定的 save_path 时，save_path 将会被其赋值，若 save_path='' 则保存在运行的py文件的目录下
        else:
            save_path = save_path

        # 判断是否添加光滑曲线
        if fitting:

            # 调用私有方法 __curve_fitting() 来获取光滑的 DataFrame 数据
            smoothing_dic, smoothing_original_precision_dic, realized_dic, realized_original_precision_dic = \
                self.__curve_fitting(degree=degree, smoothing=smoothing,
                                     noise_level=noise_level, precision=precision)

            if noise_level > 0:  # 添加噪声的情况
                result_dic = realized_dic  # 绘制曲线
                data_valid_dic = realized_original_precision_dic  # 原精度曲线
            else:  # 不添加噪声的情况
                result_dic = smoothing_dic
                data_valid_dic = smoothing_original_precision_dic

            data_black_dic = copy.deepcopy(self.data_black_dic)  # 标准曲线

        # 不添加光滑曲线的情况
        else:
            result_dic = copy.deepcopy(self.data_valid_dic)  # 绘制曲线
            data_valid_dic = copy.deepcopy(self.data_valid_dic)  # 原精度曲线
            data_black_dic = copy.deepcopy(self.data_black_dic)  # 标准曲线

        # 判断 deconvolution, area, high 是否长度相同，不同则抛出错误，只有在 deconvolution 不为 None 时才会判断
        if deconvolution is not None and not (len(deconvolution) == len(area) == len(high)):
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"deconvolution, area, and high must have the same length.")

        # 十种常用的配色方案
        color_palette = [
            '#87CEEB',  # 蓝色
            '#ff7f0e',  # 橙色
            '#2ca02c',  # 绿色
            '#d62728',  # 红色
            '#9467bd',  # 紫色
            '#8c564b',  # 棕色
            '#e377c2',  # 粉红色
            '#7f7f7f',  # 灰色
            '#bcbd22',  # 黄色
            '#17becf'  # 青色
        ]

        # 遍历 dict 以获取数据
        for title, data_df in data_valid_dic.items():

            # 提取数据的 X 和 Y 值
            x = data_df.iloc[:, 0]
            y = data_df.iloc[:, 1]

            # 初始化重构数据数组
            reconstructed_data = np.zeros_like(y, dtype=float)

            # 设置图像大小和分辨率
            plt.figure(figsize=(6, 4.5), dpi=200)

            # 初始化标准数据和标准范围
            black_data = None
            in_black_range = np.ones_like(y, dtype=bool)

            # 判断是否存在标准数据
            if data_black_dic and title in data_black_dic:
                black_data = data_black_dic[title].iloc[:, 1]
                # 获取标准数据大于 0 的范围
                in_black_range = black_data > 0

            # 判断是否进行反卷积
            if deconvolution is not None:
                # 循环处理每个反卷积的峰值
                for i, (peak_idx, peak_area, peak_high) in enumerate(zip(deconvolution, area, high)):
                    color = color_palette[i % len(color_palette)]
                    width = peak_area / peak_high

                    # 计算峰值数据
                    peak_data = peak_high * np.exp(-0.5 * ((x - peak_idx) / width) ** 2)

                    # 有标准数据的情况
                    if black_data is not None:
                        # 根据标准范围更新重构数据
                        reconstructed_data[in_black_range] += peak_data[in_black_range]
                        # 绘制填充区域
                        plt.fill_between(x, peak_data + black_data, black_data,
                                         color=color, alpha=0.3, where=in_black_range)

                    # 无标准数据的情况
                    else:
                        reconstructed_data += peak_data
                        plt.fill_between(x, peak_data, color=color, alpha=0.3, **kwargs)

                # 以光滑曲线代替反卷积拟合曲线的情况
                if show_perfect:

                    # 绘制重构的峰值数据
                    plt.plot(x[in_black_range],
                             reconstructed_data[in_black_range] +
                             (black_data[in_black_range] if black_data is not None else 0),
                             label='Reconstructed Line',
                             linestyle='-',
                             color='red')

                    # 数据整合
                    x = x[in_black_range]
                    y = reconstructed_data[in_black_range] + \
                        (black_data[in_black_range] if black_data is not None else 0)

                    # 每隔几个点绘制一次
                    step = 4
                    x_to_plot = x[::step]
                    y_to_plot = y[::step]

                    # 创建一个DataFrame
                    self.y_result_df = pd.DataFrame({'Binding energy (eV)': x_to_plot, 'Intensity (a.u.)': y_to_plot})

                    noised_dic = self.__add_noise(noise_level=noise_level)
                    noised_df = list(noised_dic.values())[0]

                    # 绘制数据
                    plt.scatter(x=noised_df.iloc[:, 0],
                                y=noised_df.iloc[:, 1],
                                facecolors='none',  # 内圈颜色
                                edgecolors='black',  # 外圈颜色
                                marker='o',
                                s=30,
                                alpha=1,
                                label="Data Subset")

                # 正常绘制的情况
                else:

                    # 绘制重构的峰值数据
                    plt.plot(x[in_black_range],
                             reconstructed_data[in_black_range] +
                             (black_data[in_black_range] if black_data is not None else 0),
                             label='Reconstructed Line',
                             linestyle='--',
                             color='red')

                    # 绘制原始数据
                    plt.scatter(x=result_dic[title].iloc[:, 0],
                                y=result_dic[title].iloc[:, 1],
                                facecolors='none',  # 内圈颜色
                                edgecolors='black',  # 外圈颜色
                                marker='o',
                                s=30,
                                alpha=1,
                                label=title)

            # 若存在标准数据，则绘制标准线
            if black_data is not None:
                plt.plot(x[in_black_range], black_data[in_black_range],
                         label='Standard Line', linestyle='--', color='grey')

            # 设置坐标轴标签和标题
            plt.xlabel(self.data.x_label, fontdict=self.data.font_title)
            plt.ylabel(self.data.y_label, fontdict=self.data.font_title)

            # 刻度轴字体
            plt.xticks(fontsize=self.data.font_ticket['size'],
                       fontweight=self.data.font_ticket['weight'],
                       fontfamily=self.data.font_ticket['family'])
            plt.yticks([])

            plt.xlim(np.min(x), np.max(x))  # 设置最大和最小值

            # 只有当 show_legend 为 True 时才会有图注
            if show_legend:
                plt.legend(prop=self.data.font_legend)
            else:
                plt.legend().remove()

            # 反转 X 轴
            plt.gca().invert_xaxis()

            # 调整布局并显示图像
            plt.tight_layout()

            # 如果提供了保存路径，则保存图像到指定路径
            if save_path is not None:  # 如果 save_path 的值不为 None，则保存
                file_name = title + ".png"  # 初始文件名为 "title.png"
                full_file_path = os.path.join(save_path, file_name)  # 创建完整的文件路径

                if os.path.exists(full_file_path):  # 查看该文件名是否存在
                    count = 1
                    file_name = title + f"_{count}.png"  # 若该文件名存在则在后面加 '_1'
                    full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                    while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                        count += 1
                        file_name = title + f"_{count}.png"
                        full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                plt.savefig(fname=full_file_path, dpi=dpi)  # 使用完整路径将散点图保存到指定的路径

            plt.show()
            time.sleep(self.interval_time)  # 让程序休息一段时间，防止绘图过快导致程序崩溃

        # 判断是否打印信息
        if show_information:

            if deconvolution is not None:  # 判断是否打印 deconvolution 等信息 (主要)
                # 计算总面积
                total_area = sum(area)

                # 初始化表格
                table = PrettyTable()
                table.field_names = ["Deconvolution", "Area", "High", "Area Percentage"]

                # 填充表格数据
                for d, a, h in zip(deconvolution, area, high):
                    area_percentage = (a / total_area) * 100  # 计算面积占比
                    table.add_row([d, a, h, f"{area_percentage:.2f}%"])

                # 打印表格
                print(table)

                if fitting:  # 判断是否打印拟合信息
                    print("The curve has been fitted.")

                if noise_level > 0:  # 判断是否打印噪声信息
                    print("Noise Level:\t", noise_level)

            else:
                print("Deconvolution is not applied.")

                if fitting:  # 判断是否打印拟合信息
                    print("The curve has been fitted.")

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 拟合
    def __curve_fitting(self, degree: int = 5, smoothing: float = 5, noise_level: float = 0, precision: float = 0.2)\
            -> Tuple[Dict[str, DataFrame], Dict[str, DataFrame], Dict[str, DataFrame], Dict[str, DataFrame]]:
        """
        进行曲线拟合，如果 add_fitting == True 则调用此函数
        Perform curve fitting; call this function if add_fitting == True.

        :param degree: (int) 曲线阶数，越大越平滑，1 <= degree_fitting <= 5
        :param smoothing: (float) 平滑因子，越小越接近原数据
        :param noise_level: (float) 添加的噪声程度，此项必需大于等于 0，默认为 0，即不添加
        :param precision: (float) 绘制散点的精度，默认为 0.02

        :return smoothing_dic: (dict) 保存光滑后曲线的 dict
        :return smoothing_original_precision_dic: (dict) 保存光滑后曲线原精度的 dict
        :return realized_dic: (dict) 保存光滑并添加噪声后曲线的 dict
        :return realized_original_precision_dic: (dict) 保存光滑并添加噪声后曲线原精度的 dict
        """

        data_dic = copy.deepcopy(self.data_valid_dic)

        # 实例化 general.Magic 来光滑曲线
        self.curve_smoothing = general.Magic(keyword='XPS')
        self.curve_smoothing.smooth_curve(data_dic=data_dic,
                                          degree_smoothing=degree,
                                          smooth_smoothing=smoothing)

        # 绘制的光滑曲线
        smoothing_dic = self.curve_smoothing.improve_precision(spline_dic=self.curve_smoothing.spline_smoothing_dic,
                                                               precision_smoothing=precision)
        self.smoothing_dic = smoothing_dic

        # 进行结合的光滑曲线
        smoothing_original_precision_dic = self.curve_smoothing.improve_precision(
            spline_dic=self.curve_smoothing.spline_smoothing_dic,
            precision_smoothing=0.05)  # 测试数据的原精度为 0.05
        self.smoothing_original_precision_dic = smoothing_original_precision_dic

        realized_dic = None
        realized_original_precision_dic = None

        # 添加噪声使得数据更加真实
        if noise_level > 0:

            # 对光滑曲线添加噪声
            realized_dic = self.curve_smoothing.realize_data(data_dic=smoothing_dic,
                                                             noise_level=noise_level,
                                                             protected_column=0)
            self.realized_dic = realized_dic

            # 对光滑曲线的原精度添加噪声
            realized_original_precision_dic = self.curve_smoothing.realize_data(
                data_dic=smoothing_original_precision_dic,
                noise_level=noise_level)
            self.realized_original_precision_dic = realized_original_precision_dic

        return smoothing_dic, smoothing_original_precision_dic, realized_dic, realized_original_precision_dic

    # 加噪
    def __add_noise(self, noise_level: float = 0) -> Dict[str, DataFrame]:
        """
        给数据添加噪声，使得数据更加真实
        Add noise to the data to make it more realistic.

        :param noise_level: (float) 添加噪声的程度

        :return noised_dic: (dict) 添加噪声后的 dict
        """

        y_result_df = self.y_result_df

        # # 实例化 general.Magic 来为曲线添加噪声
        self.curve_noising = general.Magic(keyword='XPS', data_df=y_result_df)
        self.curve_noising.data_init()  # 数据分配

        noised_dic = self.curve_noising.realize_data(data_dic=self.curve_noising.data_dic,
                                                    noise_level=noise_level,
                                                    protected_column=0)

        return noised_dic


""" Mapping """
class Mapping(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # analyze_Mapping()
        self.image_dic = None

        # __composite_image()
        self.result_image_dic = None

        # __print_details()
        self.element_dic = None

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(keyword='Mapping', save_path=save_path)

    # 读取原生文件
    def read(self, delimiter: Optional[str] = None) -> Dict[str, DataFrame]:
        r"""
        读取能谱的源生 TXT 文件
        Read the raw TXT file of the energy spectrum.

        针对仪器： SEM-EDS 乌灯丝扫描电子显微镜 SU3500，放大倍数：5~30000；样品台尺寸：装载直径 ≥ 200mm; 有 S E以及 BSE 成像

        :param delimiter: (str) TXT 文件中的分割符，用以分列，默认为 keyword = 'Mapping' 中的 self.delimiter 参数，即 ','

        :return data_dic: (dict) 数据的 dict 格式 {'title': DataFrame}
        DataFrame: column1:       float64
                   column2:       float64
                   column3:       float64
                   dtype:         object
        """

        if delimiter is not None:
            self.data.delimiter = delimiter

        self.data.read_txt(txt_path=self.read_path)  # 读取 TXT 文件数据

        data_dic = copy.deepcopy(self.data.data_dic)

        return data_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, line_color: Optional[str] = None,
             fill_color: Optional[str] = None, background_color: Optional[str] = None, **kwargs) -> None:
        """
        绘制能谱图像
        Plot the image of energy spectrum.

        :param save_path: (str) 文件的保存路径
        :param line_color: (str) 线条的颜色
        :param fill_color: (str) 填充面积的颜色
        :param background_color: (str) 背景的颜色
        :param kwargs: 可加的关键词参数

        :return: None
        """

        if line_color is not None:
            self.data.line_color = 'purple'

        if fill_color is not None:
            fill_color = fill_color
        else:
            fill_color = 'skyblue'

        if background_color is not None:
            background_color = background_color
        else:
            background_color = self.data.background_color

        self.data.plot_line(

            # 填充折线图下方的区域
            fill_area=True,
            fill_color=fill_color,
            fill_alpha=0.5,

            # 背景设置
            background_color=background_color,

            **kwargs,
        )  # 进行图形的绘制

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 分析 Mapping 的报告 word，并整合图片
    def analyze_Mapping(self, read_path: Optional[str] = None, save_path: Union[bool, str] = True,
                        composite_image: bool = True, print_detail: bool = True, dpi: int = 600) -> Dict[str, list]:
        """
        分析 SEM-EDS Mapping 得到的 Word 文档，进行合成图像，和提取并打印元素含量的功能
        Word documents obtained by SEM-EDS Mapping can be analyzed to synthesize images,
        and extract and print element content

        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        :param composite_image: (bool) 是否合成图像，默认为 True
        :param print_detail: (bool) 是否打印元素含量的详细信息，默认为 True
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600

        :return image_dic: (dict) 包含图像的 dict，一个键对应三个图像，第一个为 SEM 图像，第二个为元素分布图像，第三个为元素含量图像
        """

        # 第一部分：检查路径
        if read_path is None:
            read_path = self.read_path

        # 当 save_path==True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            save_path = self.save_path
        # 若 save_path 为 None 或 False 时，本图形不保存
        elif save_path is False:
            save_path = None
        # 当有指定的 save_path 时，save_path 将会被其赋值，若 save_path == '' 则保存在运行的 py 文件的目录下
        else:
            save_path = save_path

        word_list = []  # 创建一个空的 list 用于存储有效路径
        title_list = []  # 创建一个空的 list 用于存储标题

        # 检查路径是否存在
        if not os.path.exists(read_path):
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"The provided path does not exist.")

        # 检查路径是否是一个目录
        elif os.path.isdir(read_path):
            # 获取目录下所有文件
            files = os.listdir(read_path)
            # 过滤出所有的 Word 文件
            word_files = [f for f in files if f.endswith('.docx')]
            # 如果目录下没有 Word 文件，则报错
            if not word_files:
                class_name = self.__class__.__name__  # 获取类名
                method_name = inspect.currentframe().f_code.co_name  # 获取方法名
                raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                                 f"The directory does not contain any Word files.")
            # 打开目录下所有的 Word 文件
            for word_file in word_files:

                word_path = os.path.join(read_path, word_file)  # 获取有效路径
                title_without_extension, _ = os.path.splitext(word_file)  # 使用 splitext() 分割文件名和扩展名

                word_list.append(word_path)
                title_list.append(title_without_extension)

        # 检查路径是否是一个 Word 文件
        elif read_path.endswith('.docx'):

            file_name = re.search(r'([^/\\]+)\.docx$', read_path).group(1)  # 获取文件名

            word_list.append(read_path)
            title_list.append(file_name)

        else:
            # 如果不是目录也不是Word文件，报错
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"The path is neither a directory nor a Word file.")

        # 第二部分：处理 Word 文档
        image_dic = {}
        for i, (word_path) in enumerate(word_list):
            doc = Document(word_path)
            figures = []

            # 遍历文档中的所有段落，并提取图片
            for rel_id, rel in doc.part.rels.items():
                if "image" in rel.reltype:
                    image = rel.target_part.blob
                    image_stream = io.BytesIO(image)
                    figures.append(Image.open(image_stream))

                    # 如果我们已经找到了两个图片，就停止搜索
                    if len(figures) == 6:
                        break

            target_figures = [figures[2], figures[4], figures[5]]
            image_dic[title_list[i]] = target_figures

        self.image_dic = image_dic

        # 合成图像
        if composite_image:
            self.__composite_image(save_path=save_path, dpi=dpi)

        # 打印元素含量详细信息
        if print_detail:
            self.__print_element_details()

        return image_dic

    # 合成图像
    def __composite_image(self, save_path: Union[bool, str] = True, dpi: int = 600) -> dict:
        """
        将两个图像合成一张。将第二张图像的黑色换成透明色，并将其中的彩色加深，使其更加鲜艳，然后放入第一张图片上
        Combine two images into one. Replace the black of the second image with a transparent color,
        and deepen the color in it to make it more vivid, and then put it on the first image

        :param save_path: (str) 图像的保存路径
        :param dpi: (int) 图片保存的精度，只有在需要保存时才有意义，默认为 dpi=600

        :return result_image_dic: (dict) 包含合成图像的 dict，键为 title，值为合成后的图像
        """

        image_dic = copy.deepcopy(self.image_dic)

        result_image_dic = {}
        for title, target_figures in image_dic.items():

            image1 = target_figures[0]
            image2 = target_figures[1]

            # 色系映射表
            color_mapping = {
                "red": (255, 0, 0),
                "green": (0, 255, 0),
                "blue": (0, 0, 255),
                "yellow": (255, 255, 0)
            }

            # 将图片2的黑色和接近黑色部分变成透明色
            image2 = image2.convert("RGBA")
            data = image2.getdata()
            new_data = []

            progress_bar = tqdm(data, desc="SEM_EDS_surface_scanning", leave=False)  # 创建进度条，并在处理完成后清除进度条
            for item in progress_bar:
                # 判断像素是否接近黑色
                if item[0] < 10 and item[1] < 10 and item[2] < 10:
                    # 设置像素为完全透明
                    new_data.append((item[0], item[1], item[2], 0))
                else:
                    # 获取像素的色系
                    pixel_color = item[:3]
                    # 判断色系并进行颜色替换
                    if pixel_color[0] > pixel_color[1] and pixel_color[0] > pixel_color[2]:
                        new_data.append(color_mapping["red"] + (item[3],))
                    elif pixel_color[1] > pixel_color[0] and pixel_color[1] > pixel_color[2]:
                        new_data.append(color_mapping["green"] + (item[3],))
                    elif pixel_color[2] > pixel_color[0] and pixel_color[2] > pixel_color[1]:
                        new_data.append(color_mapping["blue"] + (item[3],))
                    else:
                        new_data.append(color_mapping["yellow"] + (item[3],))

            # 执行完毕后重置进度条
            progress_bar.close()
            # 执行完毕后回到起始位置
            tqdm.write("\033[F\033[K")

            # 创建处理后的图片对象
            image2.putdata(new_data)

            # 增强图片2的颜色使其更加鲜艳
            enhancer = ImageEnhance.Color(image2)
            enhanced_image2 = enhancer.enhance(0.5)  # 增强因子，可以根据需要进行调整

            # 将图片2整合到图片1上
            image1.paste(enhanced_image2, (0, 0), enhanced_image2)

            result_image_dic[title] = image1

            plt.figure(figsize=(6, 4.5), dpi=200)
            plt.imshow(image1)
            plt.axis('off')  # 不显示坐标轴

            # 如果提供了保存路径，则保存图像到指定路径
            if save_path is not None:  # 如果 save_path 的值不为 None，则保存
                file_name = title + ".png"  # 初始文件名为 "title.png"
                full_file_path = os.path.join(save_path, file_name)  # 创建完整的文件路径

                if os.path.exists(full_file_path):  # 查看该文件名是否存在
                    count = 1
                    file_name = title + f"_{count}.png"  # 若该文件名存在则在后面加 '_1'
                    full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                    while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                        count += 1
                        file_name = title + f"_{count}.png"
                        full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                plt.savefig(fname=full_file_path, dpi=dpi)  # 使用完整路径将散点图保存到指定的路径

            plt.show()
            time.sleep(self.interval_time)  # 让程序休息一段时间，防止绘图过快导致程序崩溃

        self.result_image_dic = result_image_dic

        return result_image_dic

    # 打印元素含量信息
    def __print_element_details(self) -> Dict[str, dict]:
        """
        读取图片中元素的含量信息并以 tabulate 表格的形式打印出来
        Read the content information of the elements in the picture and print it out in tabulate form

        :return element_dic: (dict) 包含元素含量信息的 dict，键为 title，值为元素含量信息
        """

        image_dic = copy.deepcopy(self.image_dic)

        element_dic = {}
        for title, target_figures in image_dic.items():

            # 第四部分：识别元素含量
            image = target_figures[2]

            # 使用 pytesseract 对预处理后的图片进行 OCR
            text = pytesseract.image_to_string(image)

            # 正则表达式匹配模式：数字%、空格、元素符号（不包含K），可能的小写字母（表示化学线，如L）
            pattern = re.compile(r'(\d+)%\s+([A-Z][a-z]?)(?:K|L)?')

            # 提取元素和其百分比含量
            matches = pattern.findall(text)

            # 创建字典以存储元素和对应的百分比
            element_percentage_dict = {element: percentage for percentage, element in matches}

            # 如果只显示了 'A'，那么认为是 'Al' 没有被识别准确，则将 'A' 替换为 'Al'
            for element in list(element_percentage_dict.keys()):
                if element == 'A':
                    element_percentage_dict['Al'] = element_percentage_dict.pop('A')

            element_dic[title] = element_percentage_dict

        self.element_dic = element_dic

        # 调整表格数据，使得样品名称在其元素之间垂直居中，并在不同样品间添加横线
        table_data_grouped = []
        for sample, elements in element_dic.items():
            # 计算每个样品的元素数量
            num_elements = len(elements)
            # 计算样品名称应该出现在哪一行（中间位置）
            sample_row = num_elements // 2
            for i, (element, percentage) in enumerate(elements.items()):
                # 将样品名称放在中间位置
                row_sample = sample if i == sample_row else ""
                table_data_grouped.append([row_sample, element, f"{percentage}%"])
            # 在每个样品之后添加分隔行，除了最后一个样品
            if sample != list(element_dic.keys())[-1]:
                table_data_grouped.append(["-" * 15, "-" * 10, "-" * 10])  # 使用适当数量的破折号以适应列宽

        # 使用 tabulate 打印表格
        table_grouped = tabulate(table_data_grouped, headers=["Sample", "Element", "Percentage"], tablefmt="pretty")
        print(table_grouped)

        return element_dic


""" XRF """
class XRF(Keyword):

    # 初始化
    def __init__(self, read_path: str, save_path: Optional[str] = None):
        """
        :param read_path: (str) 数据的读取路径，可是文件路径或目录路径
        :param save_path: (str) 图像的保存路径
        """

        # 数据接入类中
        if not os.path.isabs(read_path):  # 检查路径是否是绝对路径
            read_path = os.path.abspath(read_path)  # 如果是相对路径，则将其转换为绝对路径
        self.read_path = read_path
        if save_path is not None and not os.path.isabs(save_path):  # 如果路径不为 None 且为相对路径
            save_path = os.path.abspath(save_path)  # 如果是相对路径，则将其转换为绝对路径
        self.save_path = save_path

        # read()
        self.lines = None
        self.element_dic = None

        # plot()
        self.data = None

        # print_element()
        self.element_filtrated_dic = None

        # __extract_data()
        self.data_dic = None

    # 读取原生文件
    def read(self, locator: str = 'Analyte', file_pattern: str = r'.*') -> Dict[str, Dict[str, float]]:
        """
        读取 XRF 的原生 TXT 文件
        Read the raw TXT file of the XRF.

        针对仪器： 日本岛津制作所制备的 XRF ，型号为 XRF-1800

        :param locator: (str) 定位符，默认为 'Analyte'
        :param file_pattern: (str) file_pattern 为目录下文件的正则表达式匹配，只有 path 为目录时才有意义，默认为所有文件

        :return data_dic: (dict) 数据的 dict 格式 {'title': {Si: 56.74, ...}, ...}
        """

        # 检查文件是否存在且为 TXT 文件
        if os.path.isfile(self.read_path) and (self.read_path.endswith('.txt') or self.read_path.endswith('.TXT')):

            # 检查 TXT 文件的编码类型
            with open(self.read_path, 'rb') as f:
                result = chardet.detect(f.read())
                encoding = result['encoding']

            # 打开 TXT 文件
            with open(self.read_path, 'r', encoding=encoding) as f:
                lines = f.readlines()  # 读取 TXT 文件
                self.lines = lines

            # 提取标题
            title = re.match(pattern=r'\/.*\/(?P<name>.*)\..*$', string=self.read_path).group('name')

            # 提取数据
            df1 = self.__extract_data(start_range=10.0000, end_range=90.0000, step=0.1)
            df2 = self.__extract_data(start_range=90.0000, end_range=140.0000, step=0.1)
            # df3 = self.__extract_data(140.5000, 148.0000, 0.1)

            # 合并数据框
            final_df = pd.concat([df1, df2]).reset_index(drop=True)

            data_dic = {title: final_df}

            # 将行连接成单个字符串
            text = "".join(lines)

            analyte_index = text.find(locator)  # 查找从 'Analyte' 开始的文本部分
            analyte_text = text[analyte_index:]  # 仅提取从 'Analyte' 开始的文本部分

            # 正则表达式查找元素及其含量
            pattern = r"(\w+)\s+(\d+\.\d+) %"  # 修改了正则表达式以排除百分号

            # 解析数据
            matches = re.findall(pattern, analyte_text)
            element_con_dict = {match[0]: float(match[1]) for match in matches}  # 转换为浮点数

            # 返回该 TXT 文件的 dict，key 为文件名，value 为元素含量的字符串 (带 %)
            element_dic = {title: element_con_dict}

            self.data_dic = data_dic
            self.element_dic = element_dic

            return element_dic

        # 收到的路径为 TXT 文件目录路径的情况
        elif os.path.isdir(self.read_path):

            # 创造空的 dict 以存储数据
            data_dic = {}
            element_dic = {}
            for file in os.listdir(self.read_path):
                file_path = os.path.join(self.read_path, file)  # 目录下所有的文件的路径
                if re.search(file_pattern, file) and (file_path.endswith('.txt') or file_path.endswith('.TXT')):
                    # 检查 TXT 文件的编码类型
                    with open(file_path, 'rb') as f:
                        # 获取文件的编码类型
                        result = chardet.detect(f.read())
                        encoding = result['encoding']

                    with open(file_path, 'r', encoding=encoding) as f:
                        lines = f.readlines()  # 读取 TXT 文件
                        self.lines = lines

                    # 提取标题
                    title = re.match(pattern=r'\/.*\/(?P<name>.*)\..*$', string=file_path).group('name')

                    # 提取数据
                    df1 = self.__extract_data(start_range=10.0000, end_range=90.0000, step=0.1)
                    df2 = self.__extract_data(start_range=90.0000, end_range=140.0000, step=0.1)
                    # df3 = self.__extract_data(140.5000, 148.0000, 0.1)

                    # 合并数据框
                    final_df = pd.concat([df1, df2]).reset_index(drop=True)

                    data_dic[title] = final_df

                    # 将行连接成单个字符串
                    text = "".join(lines)

                    analyte_index = text.find(locator)  # 查找从 'Analyte' 开始的文本部分
                    analyte_text = text[analyte_index:]  # 仅提取从 'Analyte' 开始的文本部分

                    # 正则表达式查找元素及其含量
                    pattern = r"(\w+)\s+(\d+\.\d+) %"  # 修改了正则表达式以排除百分号

                    # 解析数据
                    matches = re.findall(pattern, analyte_text)
                    element_con_dict = {match[0]: float(match[1]) for match in matches}  # 转换为浮点数

                    # 返回该 TXT 文件的 dict，key 为文件名，value 为元素含量的字符串 (带 %)
                    element_dic[title] = element_con_dict

            self.data_dic = data_dic
            self.element_dic = element_dic

            return element_dic

    # 绘图
    def plot(self, save_path: Union[bool, str] = True, **kwargs) -> None:
        """
        绘制能谱图像
        Drawing energy spectrum images.

        :param save_path: (str) 保存目录，应为保存的目录而非具体路径
        :param kwargs: 可加的关键词参数

        :return: None
        """

        # 当 save_path == True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            save_path = self.save_path
        # 若 save_path 为 False 时，本图形不保存
        elif save_path is False:
            save_path = None
        # 当有指定的 save_path 时，save_path 将会被其赋值，若 save_path == '' 则保存在运行的 py 文件的目录下
        else:
            save_path = save_path

        # 默认参数
        palette = sns.color_palette(palette="light:#FFB6C1", as_cmap=True)
        width_height = kwargs.pop('width_height', (6.5, 4))
        line_width = kwargs.pop('line_width', 1.5)
        line_color = kwargs.pop('line_color', ['#8B008B'])
        x_min = kwargs.pop('x_min', 10)
        x_max = kwargs.pop('x_max', 140)
        y_min = kwargs.pop('y_min', 0)
        fill_area = kwargs.pop('fill_area', True)
        fill_color = kwargs.pop('fill_color', '#33B5E5')
        background_color = kwargs.pop('background_color', palette)

        # 实例化类 general.Manager 来进行数据的读取和图形的绘制  /* 该属性为实例化类，调用类属性时需要区分 */
        self.data = general.Manager(data_dic=self.data_dic, save_path=save_path)
        self.data.plot_line(width_height=width_height,
                            line_width=line_width,
                            line_color=line_color,
                            x_min=x_min,
                            x_max=x_max,
                            y_min=y_min,
                            fill_area=fill_area,
                            fill_color=fill_color,
                            background_color=background_color,
                            **kwargs)

        return None

    # 利用 Magic_Database 中的 JSON 文件绘图
    @classmethod  # 由于类方法，可使得其在不实例化情况下调用
    def magic_plot(cls, save_path: Optional[str] = None, **kwargs):
        pass

    #  存储成 JSON 文件
    def save_json(self):
        pass

    # 打印元素含量表
    def print_element(self, elements_to_print: Union[float, list[str]] = False, decimal_places: int = 2) -> PrettyTable:
        """
        打印元素含量表
        Print element content table.

        :param elements_to_print: (float / list) 打印的元素，为 False 时打印所有元素，为 True 时只打印共有元素，为 list 时只打印
                                 给定的元素，默认为 False，表示打印所有元素
        :param decimal_places: (int) 小数点的保留位数，默认为 2

        :return table: (PrettyTable) 可打印的表格
        """

        element_dic = copy.deepcopy(self.element_dic)

        # 确定要打印的元素
        if isinstance(elements_to_print, list):
            elements_set = set(elements_to_print)
        elif elements_to_print is True:
            # 找出所有样品共有的元素
            elements_set = set.intersection(*(set(sample.keys()) for sample in element_dic.values()))
        else:
            # 包括所有元素
            elements_set = set.union(*(set(sample.keys()) for sample in element_dic.values()))

        # 创建筛选后的元素字典
        self.element_filtrated_dic = {
            sample: {element: concentration for element, concentration in concentrations.items() if
                     element in elements_set} for sample, concentrations in element_dic.items()}

        # 收集所有样品中的所有独特元素和计算它们的总含量及出现次数
        total_concentrations = {}
        element_counts = {}
        for concentrations in element_dic.values():
            for element, concentration in concentrations.items():
                if element in elements_set:
                    total_concentrations[element] = total_concentrations.get(element, 0) + concentration
                    element_counts[element] = element_counts.get(element, 0) + 1

        # 计算每个元素的平均含量
        average_concentrations = {element: total / count for element, total in total_concentrations.items() for count in
                                  [element_counts[element]]}

        # 根据平均含量对元素进行排序
        sorted_elements = sorted(elements_set, key=lambda elem: average_concentrations.get(elem, 0), reverse=True)

        # 创建 PrettyTable 表格，列为排序后的元素
        table = PrettyTable()
        table.field_names = ["Sample"] + sorted_elements

        # 为每个样品添加一行，包含每个元素的含量
        for sample, concentrations in element_dic.items():
            row = [sample] + [
                f"{concentrations.get(element, ''):.{decimal_places}f} %" if element in concentrations else '' for
                element in sorted_elements]
            table.add_row(row)

        # 添加一条横线分隔符
        table.add_row(['-' * len(table.field_names)] + ['-------' for _ in sorted_elements])

        # 在表格底部添加每一列的平均值
        averages_row = ["Average"] + [f"{average_concentrations.get(element, 0):.{decimal_places}f} %" for element in
                                      sorted_elements]
        table.add_row(averages_row)

        # 打印表格
        print(table)

        return table

    # 保存到 Excel 表格中
    def save_to_excel(self, save_path: Union[bool, str] = True, decimal_places: int = 2,
                      save_sample_name: bool = False) -> None:
        """
        保存筛选后的元素含量到 Excel 表格中
        Save the filtered element content to an Excel table.

        :param save_path: (str) 保存目录，应为保存的目录而非具体路径
        :param decimal_places: (int) 小数点的保留位数，默认为 2
        :param save_sample_name: (bool) 是否保存样本名作为第一列，默认为 False

        :return: None
        """

        if self.element_filtrated_dic is not None:
            element_save_dic = self.element_filtrated_dic
        elif self.element_dic is not None:
            element_save_dic = self.element_dic
        else:
            class_name = self.__class__.__name__  # 获取类名
            method_name = inspect.currentframe().f_code.co_name  # 获取方法名
            raise ValueError(f"\033[95mIn {method_name} of {class_name}\033[0m, "
                             f"There is no element content data to save.")

        # 当 save_path == True 时，沿用 self.save_path 的设置，此项为默认项
        if save_path is True:
            save_path = self.save_path
        # 若 save_path 为 False 时，本图形不保存
        elif save_path is False:
            save_path = None
        # 当有指定的 save_path 时，line_save_path 将会被其赋值，若 save_path == '' 则保存在运行的py文件的目录下
        else:
            save_path = save_path

        # 确定列名
        columns = ["Sample"] if save_sample_name else []
        # 使用字典中的第一个样本来获取元素名称的顺序
        elements = list(next(iter(element_save_dic.values())).keys())
        columns += elements

        # 将数据转换为 DataFrame
        data = []
        for sample, concentrations in element_save_dic.items():
            row = [sample] if save_sample_name else []
            for element in elements:
                if element in concentrations:
                    row.append(round(concentrations[element], decimal_places))
                else:
                    row.append(0)
            data.append(row)

        df = pd.DataFrame(data, columns=columns)

        # 保存至 Excel
        if save_path is not None:  # 如果 save_path 的值不为 None，则保存
            file_name = "XRF" + ".xlsx"  # 初始文件名为 XRF.xlsx
            full_file_path = os.path.join(save_path, file_name)  # 创建完整的文件路径

            if os.path.exists(full_file_path):  # 查看该文件名是否存在
                count = 1
                file_name = "XRF" + f"_{count}.xlsx"  # 若该文件名存在则在后面加 '_1'
                full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

                while os.path.exists(full_file_path):  # 找是否存在，并不断 +1，直到不重复
                    count += 1
                    file_name = "XRF" + f"_{count}.xlsx"
                    full_file_path = os.path.join(save_path, file_name)  # 更新完整的文件路径

            # 使用 ExcelWriter 保存 DataFrame
            with pd.ExcelWriter(full_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

                # 获取 workbook 和 sheet 以进行格式化
                worksheet = writer.sheets['Sheet1']

                # 设置单元格居中
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                               max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 设置列宽
                for col in worksheet.columns:
                    worksheet.column_dimensions[col[0].column_letter].width = 15

                # 设置第一行的背景色为淡蓝色
                for cell in worksheet[1]:
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

                # 设置其他行的背景色为绿色
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

            # 成功保存后打印
            print(f"File saved successfully at \033[92m{full_file_path}\033[0m")

        return None

    # 寻找有效数据
    def __extract_data(self, start_range: float, end_range: float, step: float) -> DataFrame:
        """
        从文件内容中提取有效数据
        Extract valid data from file content

        :param start_range: (float) 开始的 Range
        :param end_range: (float) 结束的 Range
        :param step: (float) Range 的步长

        :return data_part_df: (DataFrame) 提取到的片段数据表格
        """

        data = []  # 初始化一个空列表来存储数据
        start_collecting = False  # 设置一个标志，用于开始收集数据的条件
        locator_line = f"Range        : {start_range:.4f} - {end_range:.4f}"  # 格式化的范围字符串
        current_range = start_range  # 初始化当前范围

        for line in self.lines:
            if locator_line in line:
                # 当找到含有特定范围的行时，设置标志为True，开始收集数据
                start_collecting = True
                continue

            if start_collecting:
                # 检查是否到达了数据部分的结束位置
                if 'Channel_info_end' in line:
                    break  # 遇到'Channel_info_end'时停止读取
                else:
                    try:
                        # 检查行是否以数字开头，确保只收集数据行
                        if line.strip()[0].isdigit():
                            # 将行转换为浮点数并添加到数据列表
                            data.append([current_range, float(line.strip())])
                            current_range += step  # 更新当前范围
                    except ValueError:
                        pass  # 如果转换失败，则跳过该行

        # 创建DataFrame并返回
        data_part_df = pd.DataFrame(data, columns=["Range", "Intensity"])

        return data_part_df


""" 应用 Duck Typing 来调用实例中的绘图方法 """
def plot(keyword, save_path: Optional[str] = None, **kwargs) -> None:
    """
    利用已实例化的关键词对象进行快速绘图
    Quick plotting using instantiated keyword objects.

    :param keyword: (object) 已实例化的关键词对象
    :param save_path: (str) 保存路径
    :param kwargs: 关键词参数，用于修改 plot 方法中的属性

    :return: None
    """

    keyword.plot(save_path=save_path, **kwargs)

    return None
